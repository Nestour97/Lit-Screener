"""
Excel output writer.
Produces extraction_output.xlsx with sheets:
  Input | Classification | Extraction | Log | Summary
"""

import logging
from pathlib import Path
from typing import List, Dict, Any, Optional
from datetime import datetime

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models.schemas import ClassificationResult, ExtractionResult, LogEntry

logger = logging.getLogger(__name__)

# Colour palette for review flags
RED_FILL = PatternFill("solid", fgColor="FFCCCC")
YELLOW_FILL = PatternFill("solid", fgColor="FFFFCC")
HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(color="FFFFFF", bold=True)

MODEL_TYPE_LABELS = {1:"OLS",2:"Logit/Probit",3:"FE",4:"RE",5:"GLS",
                     6:"GMM",7:"HLM",8:"SEM",9:"Count",10:"Heckman",11:"Other"}
MISSING_LABELS = {1:"Listwise",2:"Pairwise",3:"Mean sub.",4:"Regression imp.",
                  5:"MI",6:"FIML",7:"EM",8:"Hot-deck",9:"NR",10:"Other"}
DS_LABELS = {1:"Survey",2:"Archival",3:"Content",4:"Experiment",5:"Other"}
VT_LABELS = {1:"Continuous",2:"Binary",3:"Count",4:"Ordinal",5:"Categorical"}
UA_LABELS = {1:"Individual",2:"Team",3:"Firm",4:"Industry",5:"Country",6:"Dyad",7:"Other"}
DT_LABELS = {1:"Cross-sectional",2:"Panel",3:"Time-series",4:"Mixed"}
REPL_LABELS = {1:"High",2:"Medium",3:"Low",4:"Not feasible"}
DIR_LABELS = {1:"Positive",2:"Negative",3:"Curvilinear",4:"No direction"}


def _nr_na(v, applicable=True):
    if v is None or v == "":
        return "NA" if not applicable else "NR"
    return str(v)


def write_workbook(
    input_files: List[str],
    classifications: List[ClassificationResult],
    extractions: List[ExtractionResult],
    log_entries: List[LogEntry],
    summary_md: Optional[str],
    output_path: Path,
    stats: Optional[Dict[str, Any]] = None,
) -> Path:
    """Write the final Excel workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(str(output_path), engine="openpyxl") as writer:
        _write_input_sheet(writer, input_files, classifications)
        _write_classification_sheet(writer, classifications)
        _write_extraction_sheet(writer, extractions)
        _write_log_sheet(writer, log_entries)
        _write_summary_sheet(writer, summary_md, stats, classifications, extractions)

    # Post-process: add formatting
    _format_workbook(output_path, extractions, log_entries)
    logger.info(f"[excel_writer] Workbook written: {output_path}")
    return output_path


def _write_input_sheet(writer, input_files, classifications):
    cls_map = {c.paper_id: c for c in classifications}
    rows = []
    for f in input_files:
        pid = Path(f).stem
        cls = cls_map.get(pid)
        rows.append({
            "Paper_ID": pid,
            "Filename": Path(f).name,
            "Status": "processed" if cls else "pending",
            "Classification": cls.classification if cls else "",
            "Eligible": "Yes" if (cls and cls.eligible_for_full_extraction) else "No",
            "Needs_Review": "Yes" if (cls and cls.needs_review) else "No",
        })
    pd.DataFrame(rows).to_excel(writer, sheet_name="Input", index=False)


def _write_classification_sheet(writer, classifications):
    rows = []
    for c in classifications:
        rows.append({
            "Paper_ID": c.paper_id,
            "Title": c.title,
            "Authors": c.authors,
            "Year": c.year or "NR",
            "Journal": c.journal or "NR",
            "Classification": c.classification,
            "Eligible": "Yes" if c.eligible_for_full_extraction else "No",
            "Confidence": c.confidence,
            "Needs_Review": "Yes" if c.needs_review else "No",
            "Flag_Reason": _nr_na(c.flag_reason, applicable=False),
            "Rationale": c.rationale,
        })
    pd.DataFrame(rows).to_excel(writer, sheet_name="Classification", index=False)


def _write_extraction_sheet(writer, extractions):
    rows = []
    for e in extractions:
        rows.append({
            "Paper_ID": str(e.paper_id),
            "Base_Paper_ID": str(e.base_paper_id),
            "Authors": e.authors,
            "Year": e.year or "NR",
            "Journal": e.journal or "NR",
            "Title": e.title,
            # Cat 2
            "RQ_Summary": _nr_na(e.rq_summary),
            "Num_Hypotheses": _nr_na(e.num_hypotheses),
            "Primary_Relationship": _nr_na(e.primary_relationship),
            "Relationship_Direction": DIR_LABELS.get(e.relationship_direction, _nr_na(e.relationship_direction)),
            # DV
            "DV_Name": _nr_na(e.dv_name),
            "DV_Construct": _nr_na(e.dv_construct),
            "DV_Measurement": _nr_na(e.dv_measurement),
            "DV_Measurement_Page": _nr_na(e.dv_measurement_page, applicable=False),
            "DV_Source": DS_LABELS.get(e.dv_source, _nr_na(e.dv_source)),
            "DV_Type": VT_LABELS.get(e.dv_type, _nr_na(e.dv_type)),
            "DV_Num": _nr_na(e.dv_num),
            # IV
            "IV_Name": _nr_na(e.iv_name),
            "IV_Construct": _nr_na(e.iv_construct),
            "IV_Measurement": _nr_na(e.iv_measurement),
            "IV_Measurement_Page": _nr_na(e.iv_measurement_page, applicable=False),
            "IV_Source": DS_LABELS.get(e.iv_source, _nr_na(e.iv_source)),
            "IV_Type": VT_LABELS.get(e.iv_type, _nr_na(e.iv_type)),
            "IV_Num": _nr_na(e.iv_num),
            # Mediation
            "Mediator_Present": e.mediator_present,
            "Mediator_Name": "NA" if not e.mediator_present else _nr_na(e.mediator_name),
            "Mediator_Construct": "NA" if not e.mediator_present else _nr_na(e.mediator_construct),
            "Mediator_Measurement": "NA" if not e.mediator_present else _nr_na(e.mediator_measurement),
            "Mediation_Method": "NA" if not e.mediator_present else _nr_na(e.mediation_method),
            # Moderation
            "Moderator_Present": e.moderator_present,
            "Moderator_Name": "NA" if not e.moderator_present else _nr_na(e.moderator_name),
            "Moderator_Construct": "NA" if not e.moderator_present else _nr_na(e.moderator_construct),
            "Moderator_Measurement": "NA" if not e.moderator_present else _nr_na(e.moderator_measurement),
            "Moderation_Method": "NA" if not e.moderator_present else _nr_na(e.moderation_method),
            # Controls
            "Control_Num": _nr_na(e.control_num),
            "Control_List": _nr_na(e.control_list),
            "Control_Justified": _nr_na(e.control_justified),
            # Sample
            "Sample_Size": _nr_na(e.sample_size),
            "Sample_Context": _nr_na(e.sample_context),
            "Data_Type": DT_LABELS.get(e.data_type, _nr_na(e.data_type)),
            "Data_Source_Primary": _nr_na(e.data_source_primary),
            "Unit_of_Analysis": UA_LABELS.get(e.unit_of_analysis, _nr_na(e.unit_of_analysis)),
            "Time_Period": _nr_na(e.time_period),
            # Model
            "Model_Type": MODEL_TYPE_LABELS.get(e.model_type, _nr_na(e.model_type)),
            "Model_Type_Other": _nr_na(e.model_type_other, applicable=False),
            "Endogeneity_Addressed": e.endogeneity_addressed,
            "Endogeneity_Method": _nr_na(e.endogeneity_method, applicable=not e.endogeneity_addressed),
            "Robustness_Checks": _nr_na(e.robustness_checks),
            # Missing data
            "Missing_Mentioned": e.missing_mentioned,
            "Missing_Rate_Reported": e.missing_rate_reported,
            "Missing_Rate_Value": "NA" if not e.missing_rate_reported else _nr_na(e.missing_rate_value),
            "Missing_Variables": _nr_na(e.missing_variables, applicable=bool(e.missing_mentioned)),
            "Missing_Handling": MISSING_LABELS.get(e.missing_handling, _nr_na(e.missing_handling)),
            "Missing_Handling_Other": _nr_na(e.missing_handling_other, applicable=False),
            "Missing_Handling_Page": _nr_na(e.missing_handling_page, applicable=False),
            "Missing_Justified": e.missing_justified,
            "Missing_Pattern_Tested": e.missing_pattern_tested,
            "Missing_Pattern_Result": "NA" if not e.missing_pattern_tested else _nr_na(e.missing_pattern_result),
            "Missing_Sensitivity": e.missing_sensitivity,
            # Replication
            "Data_Available": e.data_available,
            "Code_Available": e.code_available,
            "Software_Reported": e.software_reported,
            "Software_Name": "NA" if not e.software_reported else _nr_na(e.software_name),
            "Replication_Feasibility": REPL_LABELS.get(e.replication_feasibility, _nr_na(e.replication_feasibility)),
            # QC
            "Confidence": e.confidence,
            "Needs_Review": "Yes" if e.needs_review else "No",
            "Flag_Reason": _nr_na(e.flag_reason, applicable=False),
            "Coding_Notes": _nr_na(e.coding_notes, applicable=False),
        })
    pd.DataFrame(rows).to_excel(writer, sheet_name="Extraction", index=False)


def _write_log_sheet(writer, log_entries):
    rows = [e.model_dump() for e in log_entries]
    pd.DataFrame(rows).to_excel(writer, sheet_name="Log", index=False)


def _write_summary_sheet(writer, summary_md, stats, classifications, extractions):
    rows = []
    if stats:
        rows.append({"Metric": "Total Papers", "Value": stats.get("total_papers", "")})
        rows.append({"Metric": "Eligible Papers", "Value": stats.get("eligible_count", "")})
        rows.append({"Metric": "Needs Review", "Value": stats.get("needs_review_count", "")})
        rows.append({"Metric": "Extractions Completed", "Value": stats.get("extractions_count", "")})
        rows.append({"Metric": "Missing Mentioned %", "Value": stats.get("missing_mentioned_pct", "")})
        rows.append({"Metric": "Missing Rate Reported %", "Value": stats.get("missing_rate_reported_pct", "")})
        rows.append({"Metric": "Missing Justified %", "Value": stats.get("missing_justified_pct", "")})
        rows.append({"Metric": "", "Value": ""})
        rows.append({"Metric": "--- By Classification ---", "Value": ""})
        for cat, cnt in (stats.get("by_category") or {}).items():
            rows.append({"Metric": cat, "Value": cnt})
        rows.append({"Metric": "", "Value": ""})
        rows.append({"Metric": "--- Model Types ---", "Value": ""})
        for m, cnt in (stats.get("model_type_distribution") or {}).items():
            rows.append({"Metric": m, "Value": cnt})
        rows.append({"Metric": "", "Value": ""})
        rows.append({"Metric": "--- Missing Handling Methods ---", "Value": ""})
        for m, cnt in (stats.get("missing_handling_distribution") or {}).items():
            rows.append({"Metric": m, "Value": cnt})

    if summary_md:
        rows.append({"Metric": "", "Value": ""})
        rows.append({"Metric": "=== SUMMARY REPORT ===", "Value": ""})
        rows.append({"Metric": summary_md, "Value": ""})

    pd.DataFrame(rows).to_excel(writer, sheet_name="Summary", index=False)


def _format_workbook(path: Path, extractions, log_entries):
    """Apply visual formatting to all sheets."""
    wb = openpyxl.load_workbook(str(path))
    for ws in wb.worksheets:
        # Header row formatting
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        # Auto-width (capped at 60)
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 60)
        ws.freeze_panes = "A2"

    # Highlight Needs_Review rows in Extraction sheet
    if "Extraction" in wb.sheetnames:
        ws = wb["Extraction"]
        headers = [c.value for c in ws[1]]
        if "Needs_Review" in headers:
            nr_col = headers.index("Needs_Review") + 1
            for row in ws.iter_rows(min_row=2):
                if row[nr_col - 1].value == "Yes":
                    for cell in row:
                        cell.fill = YELLOW_FILL

    # Highlight flag/error rows in Log sheet
    if "Log" in wb.sheetnames:
        ws = wb["Log"]
        headers = [c.value for c in ws[1]]
        if "level" in headers:
            lvl_col = headers.index("level") + 1
            for row in ws.iter_rows(min_row=2):
                lvl = row[lvl_col - 1].value
                if lvl in ("flag", "error"):
                    for cell in row:
                        cell.fill = RED_FILL if lvl == "error" else YELLOW_FILL

    wb.save(str(path))
